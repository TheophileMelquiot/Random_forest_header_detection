#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
import numpy as np

def max_consecutive_empty(values):
    max_count = 0
    current = 0
    for v in values:
        if v is None:
            current += 1
            max_count = max(max_count, current)
        else:
            current = 0
    return max_count

def compute_basic_ratios(row):
    values = [cell.value for cell in row]
    non_empty = [v for v in values if v is not None]

    if not non_empty:
        return 0, 0

    str_ratio = sum(isinstance(v, str) for v in non_empty) / len(non_empty)
    num_ratio = sum(isinstance(v, (int, float)) for v in non_empty) / len(non_empty)

    return str_ratio, num_ratio

def extract_row_features_from_row(row, total_cols, ws, row_idx):

    values = [cell.value for cell in row]
    non_empty = [v for v in values if v is not None]

    if not non_empty:
        return None

    fill_ratio = len(non_empty) / total_cols
    str_ratio = sum(isinstance(v, str) for v in non_empty) / len(non_empty)
    num_ratio = sum(isinstance(v, (int, float)) for v in non_empty) / len(non_empty)

    bold_ratio = sum(1 for c in row if getattr(c, "font", None) and c.font.bold) / total_cols
    colored_ratio = sum(
        1 for c in row
        if getattr(c, "fill", None)
        and c.fill.fgColor
        and c.fill.fgColor.rgb != "00000000"
    ) / total_cols

    row_position = row_idx / ws.max_row

    str_lengths = [len(str(v)) for v in non_empty if isinstance(v, str)]
    avg_str_len = np.mean(str_lengths or [0])
    std_str_len = np.std(str_lengths or [0])

    header_keywords = ["date", "name", "amount", "total", "id", "ref"]
    keyword_hits = sum(
        any(k in str(v).lower() for k in header_keywords)
        for v in non_empty if isinstance(v, str)
    )
    keyword_ratio = keyword_hits / len(non_empty)
    # === NOUVELLES FEATURES ===

    max_empty_streak = max_consecutive_empty(values) / total_cols

    unique_ratio = len(set(non_empty)) / len(non_empty)

    upper_ratio = sum(
        str(v).isupper() for v in non_empty if isinstance(v, str)
    ) / len(non_empty)

    special_char_ratio = sum(
        any(c in str(v) for c in ",.;:()") 
        for v in non_empty if isinstance(v, str)
    ) / len(non_empty)

    num_to_str_ratio = num_ratio - str_ratio

    num_to_str_ratio = num_ratio - str_ratio

    # ===== DIFFERENCE AVEC LIGNE SUIVANTE =====

    if row_idx < ws.max_row:
        next_row = list(ws.iter_rows(min_row=row_idx+1, max_row=row_idx+1))[0]
        next_str_ratio, next_num_ratio = compute_basic_ratios(next_row)

        delta_str_ratio = str_ratio - next_str_ratio
        delta_num_ratio = num_ratio - next_num_ratio
    else:
        delta_str_ratio = 0
        delta_num_ratio = 0

    return [
        fill_ratio,
        str_ratio,
        num_ratio,
        bold_ratio,
        colored_ratio,
        row_position,
        avg_str_len,
        std_str_len,
        keyword_ratio,
        max_empty_streak,
        unique_ratio,
        upper_ratio,
        special_char_ratio,
        num_to_str_ratio,
        delta_str_ratio,
        delta_num_ratio   ]


# In[2]:


import os
import json
import pandas as pd 
import warnings

warnings.filterwarnings("ignore", module="openpyxl")

def build_training_data(excel_folder, labels_dict):
    X, y, groups = [], [], []

    # Get real files that physically exist
    available_files = os.listdir(excel_folder)
    print("Total fichiers JSON:", len(labels_dict))
    print("Fichiers réels:", len(available_files))

    for json_name, sheets_info in labels_dict.items():

        # Try to match JSON filename with real filename
        matched_file = None
        for real_file in available_files:
            if real_file.strip() == json_name.strip():
                matched_file = real_file
                break

        if not matched_file:
            print("⚠ File not found:", json_name)
            continue

        filepath = os.path.join(excel_folder, matched_file)

        try:
                wb = openpyxl.load_workbook(filepath,data_only=True,read_only=True)
                
        except Exception as e:
            print("❌ Error opening:", matched_file, "|", e)
            continue

        for sheet_name, true_header_row in sheets_info.items():

            if sheet_name not in wb.sheetnames:
                print("⚠ Sheet not found:", sheet_name, "in", matched_file)
                continue

            ws = wb[sheet_name]

            total_cols = ws.max_column
            max_rows_to_scan = min(ws.max_row, 20)
            sheet_id = f"{matched_file}__{sheet_name}"

            for row_idx, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=max_rows_to_scan),
                    start=1):

                features = extract_row_features_from_row(row, total_cols, ws, row_idx)

                if features:
                    label = 1 if row_idx == true_header_row else 0
                    X.append(features)
                    y.append(label)
                    groups.append(sheet_id)


    print("Total samples:", len(X))

    return np.array(X), np.array(y), np.array(groups)


# In[ ]:

import lightgbm as lgb
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report
import joblib
from tqdm import tqdm as tqdm_std
import warnings
from sklearn.metrics import (
    roc_auc_score,
    average_precision_score,
    confusion_matrix,
    classification_report,
    RocCurveDisplay,
    PrecisionRecallDisplay
)
import matplotlib.pyplot as plt
import numpy as np
from sklearn.model_selection import GroupShuffleSplit
from sklearn.model_selection import GroupKFold

warnings.filterwarnings(
    "ignore",
    message='class_weight presets "balanced" or "balanced_subsample" are not recommended for warm_start'
)

# Load your labels (you create this JSON once manually)
with open("C:/Users/label.json", encoding="utf-8") as f:
    labels_dict = json.load(f)  # {"file1.xlsx": 3, "file2.xlsx": 1, ...}

X, y, groups = build_training_data("C:/Users/excel_file", labels_dict)

gkf = GroupKFold(n_splits=5)

cv_results = []

for fold, (train_idx, test_idx) in enumerate(gkf.split(X, y, groups)):
    print(f"\n================ FOLD {fold+1} ================\n")

    X_train, X_test = X[train_idx], X[test_idx]
    y_train, y_test = y[train_idx], y[test_idx]
    groups_train, groups_test = groups[train_idx], groups[test_idx]

    model = lgb.LGBMClassifier(
        n_estimators=400,
        learning_rate=0.05,
        num_leaves=20,
        max_depth=6,
        min_child_samples=20,
        subsample=0.8,
        colsample_bytree=0.8,
        random_state=42
    )

    model.fit(X_train, y_train)

    y_proba = model.predict_proba(X_test)[:, 1]

    test_df = pd.DataFrame({
        "proba": y_proba,
        "true": y_test,
        "group": groups_test
    })

    threshold = 0.35
    idx_max = test_df.groupby("group")["proba"].idxmax()

    test_df["pred"] = 0
    for idx in idx_max:
        if test_df.loc[idx, "proba"] > threshold:
            test_df.loc[idx, "pred"] = 1

    y_pred = test_df["pred"].values

    roc_auc = roc_auc_score(y_test, y_proba)
    pr_auc = average_precision_score(y_test, y_proba)

    print(classification_report(y_test, y_pred))
    print(f"ROC-AUC: {roc_auc:.4f}")
    print(f"PR-AUC: {pr_auc:.4f}")

    cv_results.append({
        "fold": fold+1,
        "roc_auc": roc_auc,
        "pr_auc": pr_auc
    })

print("\n==== CROSS-VALIDATION SUMMARY ====")
print(pd.DataFrame(cv_results))

X_train, X_test = X[train_idx], X[test_idx]
y_train, y_test = y[train_idx], y[test_idx]
groups_train, groups_test = groups[train_idx], groups[test_idx]

final_model = lgb.LGBMClassifier(
    n_estimators=400,
    learning_rate=0.05,
    num_leaves=20,
    max_depth=6,
    min_child_samples=20,
    subsample=0.8,
    colsample_bytree=0.8,
    random_state=42
)

final_model.fit(
    X_train,
    y_train,
    eval_set=[(X_test, y_test)],
    eval_metric="auc")

with open("C:/Users/new_files/header_detection_output_model.json", encoding="utf-8") as f:
    labels_dict_test = json.load(f)  # {"file1.xlsx": 3, "file2.xlsx": 1, ...}

X_ext, y_ext, groups_ext = build_training_data(
    "C:/Users/excel_external_test/",
    labels_dict_test
)

y_ext_proba = final_model.predict_proba(X_ext)[:, 1]

ext_df = pd.DataFrame({
    "proba": y_ext_proba,
    "true": y_ext,
    "group": groups_ext
})

threshold = 0.35
idx_max = ext_df.groupby("group")["proba"].idxmax()

ext_df["pred"] = 0
for idx in idx_max:
    if ext_df.loc[idx, "proba"] > threshold:
        ext_df.loc[idx, "pred"] = 1

y_ext_pred = ext_df["pred"].values

print("\n===== EXTERNAL VALIDATION RESULTS =====")
print(classification_report(y_ext, y_ext_pred))
print("ROC-AUC:", roc_auc_score(y_ext, y_ext_proba))
print("PR-AUC:", average_precision_score(y_ext, y_ext_proba))


# ===============================
# Probabilités et prédictions
# ===============================
y_proba = model.predict_proba(X_test)[:, 1]

test_df = pd.DataFrame({
    "proba": y_proba,
    "true": y_test,
    "group": groups_test  # correspond aux indices test
})

threshold = 0.35

idx_max = test_df.groupby("group")["proba"].idxmax()

test_df["pred"] = 0

for idx in idx_max:
    if test_df.loc[idx, "proba"] > threshold:
        test_df.loc[idx, "pred"] = 1


y_pred = test_df["pred"].values
fn_indices = np.where((y_test == 1) & (y_pred == 0))[0]
print("Indices faux négatifs :", fn_indices)


feature_names = [
    "fill_ratio",
    "str_ratio",
    "num_ratio",
    "bold_ratio",
    "colored_ratio",
    "row_position",
    "avg_str_len",
    "std_str_len",
    "keyword_ratio",
    "max_empty_streak",
    "unique_ratio",
    "upper_ratio",
    "special_char_ratio",
    "num_to_str_ratio",
    "delta_str_ratio",
    "delta_num_ratio",]

importance_df = pd.DataFrame({
    "feature": feature_names,
    "importance": model.feature_importances_
}).sort_values(by="importance", ascending=False)



# ===============================
# 1️⃣ CLASSIFICATION REPORT (EN PREMIER)
# ===============================
print("\n=== CLASSIFICATION REPORT ===\n")
print(classification_report(y_test, y_pred))

# ===============================
# 2️⃣ ROC-AUC
# ===============================
roc_auc = roc_auc_score(y_test, y_proba)
print(f"\nROC-AUC: {roc_auc:.4f}")

# ===============================
# 3️⃣ PR-AUC
# ===============================
pr_auc = average_precision_score(y_test, y_proba)
print(f"PR-AUC: {pr_auc:.4f}")

# ===============================
# 4️⃣ Confusion Matrix
# ===============================
cm = confusion_matrix(y_test, y_pred)

print("\nConfusion Matrix:")
print(cm)

tn, fp, fn, tp = cm.ravel()

print("\nDétails :")
print(f"True Negatives  (TN): {tn}")
print(f"False Positives (FP): {fp}")
print(f"False Negatives (FN): {fn}")
print(f"True Positives  (TP): {tp}")

print("\nFeature Importances:\n")
print(importance_df)

# ===============================
# ROC + Precision-Recall in ONE FIGURE
# ===============================

fig, axes = plt.subplots(1, 2, figsize=(12, 5))

# ROC Curve
RocCurveDisplay.from_predictions(
    y_test,
    y_proba,
    ax=axes[0]
)
axes[0].set_title("ROC Curve")

# Precision-Recall Curve
PrecisionRecallDisplay.from_predictions(
    y_test,
    y_proba,
    ax=axes[1]
)
axes[1].set_title("Precision-Recall Curve")

plt.tight_layout()
plt.show()

# Save the model
joblib.dump(model, "header_detector.pkl")


# %%

