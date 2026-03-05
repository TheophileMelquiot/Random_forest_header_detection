#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
import numpy as np

def max_consecutive_empty(values):
    max_count = 0
    current = 0
    for v in values:
        if v is None:
            current += 1
            max_count = max(max_count, current)
        else:
            current = 0
    return max_count

def extract_row_features_from_row(row, total_cols, ws, row_idx):

    values = [cell.value for cell in row]
    non_empty = [v for v in values if v is not None]

    if not non_empty:
        return None

    fill_ratio = len(non_empty) / total_cols
    str_ratio = sum(isinstance(v, str) for v in non_empty) / len(non_empty)
    num_ratio = sum(isinstance(v, (int, float)) for v in non_empty) / len(non_empty)

    bold_ratio = sum(1 for c in row if getattr(c, "font", None) and c.font.bold) / total_cols
    colored_ratio = sum(
        1 for c in row
        if getattr(c, "fill", None)
        and c.fill.fgColor
        and c.fill.fgColor.rgb != "00000000"
    ) / total_cols

    row_position = row_idx / ws.max_row

    str_lengths = [len(str(v)) for v in non_empty if isinstance(v, str)]
    avg_str_len = np.mean(str_lengths or [0])
    std_str_len = np.std(str_lengths or [0])

    header_keywords = ["date", "name", "amount", "total", "id", "ref"]
    keyword_hits = sum(
        any(k in str(v).lower() for k in header_keywords)
        for v in non_empty if isinstance(v, str)
    )
    keyword_ratio = keyword_hits / len(non_empty)
    # === NOUVELLES FEATURES ===

    max_empty_streak = max_consecutive_empty(values) / total_cols

    unique_ratio = len(set(non_empty)) / len(non_empty)

    upper_ratio = sum(
        str(v).isupper() for v in non_empty if isinstance(v, str)
    ) / len(non_empty)

    special_char_ratio = sum(
        any(c in str(v) for c in ",.;:()") 
        for v in non_empty if isinstance(v, str)
    ) / len(non_empty)

    num_to_str_ratio = num_ratio - str_ratio

    return [
        fill_ratio,
        str_ratio,
        num_ratio,
        bold_ratio,
        colored_ratio,
        row_position,
        avg_str_len,
        std_str_len,
        keyword_ratio,
        max_empty_streak,
        unique_ratio,
        upper_ratio,
        special_char_ratio,
        num_to_str_ratio
    ]


# In[2]:


import os
import json
import pandas as pd 
import warnings

warnings.filterwarnings("ignore", module="openpyxl")

def build_training_data(excel_folder, labels_dict):
    X, y = [], []

    # Get real files that physically exist
    available_files = os.listdir(excel_folder)
    print("Total fichiers JSON:", len(labels_dict))
    print("Fichiers réels:", len(available_files))

    for json_name, sheets_info in labels_dict.items():

        # Try to match JSON filename with real filename
        matched_file = None
        for real_file in available_files:
            if real_file.strip() == json_name.strip():
                matched_file = real_file
                break

        if not matched_file:
            print("⚠ File not found:", json_name)
            continue

        filepath = os.path.join(excel_folder, matched_file)

        try:
                wb = openpyxl.load_workbook(filepath,data_only=True,read_only=True)
                
        except Exception as e:
            print("❌ Error opening:", matched_file, "|", e)
            continue

        for sheet_name, true_header_row in sheets_info.items():

            if sheet_name not in wb.sheetnames:
                print("⚠ Sheet not found:", sheet_name, "in", matched_file)
                continue

            ws = wb[sheet_name]

            total_cols = ws.max_column
            max_rows_to_scan = min(ws.max_row, 50)

            for row_idx, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=max_rows_to_scan),
                    start=1):

                features = extract_row_features_from_row(row, total_cols, ws, row_idx)

                if features:
                    label = 1 if row_idx == true_header_row else 0
                    X.append(features)
                    y.append(label)


    print("Total samples:", len(X))

    return np.array(X), np.array(y)


# In[ ]:


from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report
import joblib
from tqdm import tqdm as tqdm_std
import warnings
from sklearn.metrics import (
    roc_auc_score,
    average_precision_score,
    confusion_matrix,
    classification_report,
    RocCurveDisplay,
    PrecisionRecallDisplay
)
import matplotlib.pyplot as plt
import numpy as np

warnings.filterwarnings(
    "ignore",
    message='class_weight presets "balanced" or "balanced_subsample" are not recommended for warm_start'
)

# Load your labels (you create this JSON once manually)
with open("C:/Users/new_files/label.json", encoding="utf-8") as f:
    labels_dict = json.load(f)  # {"file1.xlsx": 3, "file2.xlsx": 1, ...}

X, y = build_training_data("C:/Users/new_files/excel_file", labels_dict)
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)


n_estimators = 300

model = RandomForestClassifier(
    n_estimators=300,          # start from 0
    warm_start=True,         # VERY IMPORTANT
    n_jobs=1,                # keep safe for memory
    max_depth=50,
    class_weight="balanced"
)

for i in tqdm_std(range(n_estimators), desc="Entrainement du modèle"):
    model.n_estimators = i + 1
    model.fit(X_train, y_train)

# ===============================
# Probabilités et prédictions
# ===============================
y_proba = model.predict_proba(X_test)[:, 1]
y_pred = (y_proba > 0.3).astype(int)

fn_indices = np.where((y_test == 1) & (y_pred == 0))[0]
print("Indices faux négatifs :", fn_indices)


feature_names = [
    "fill_ratio",
    "str_ratio",
    "num_ratio",
    "bold_ratio",
    "colored_ratio",
    "row_position",
    "avg_str_len",
    "std_str_len",
    "keyword_ratio",
    "max_empty_streak",
    "unique_ratio",
    "upper_ratio",
    "special_char_ratio",
    "num_to_str_ratio"
]

importance_df = pd.DataFrame({
    "feature": feature_names,
    "importance": model.feature_importances_
}).sort_values(by="importance", ascending=False)



# ===============================
# 1️⃣ CLASSIFICATION REPORT (EN PREMIER)
# ===============================
print("\n=== CLASSIFICATION REPORT ===\n")
print(classification_report(y_test, y_pred))

# ===============================
# 2️⃣ ROC-AUC
# ===============================
roc_auc = roc_auc_score(y_test, y_proba)
print(f"\nROC-AUC: {roc_auc:.4f}")

# ===============================
# 3️⃣ PR-AUC
# ===============================
pr_auc = average_precision_score(y_test, y_proba)
print(f"PR-AUC: {pr_auc:.4f}")

# ===============================
# 4️⃣ Confusion Matrix
# ===============================
cm = confusion_matrix(y_test, y_pred)

print("\nConfusion Matrix:")
print(cm)

tn, fp, fn, tp = cm.ravel()

print("\nDétails :")
print(f"True Negatives  (TN): {tn}")
print(f"False Positives (FP): {fp}")
print(f"False Negatives (FN): {fn}")
print(f"True Positives  (TP): {tp}")

importance_df.plot(
    kind="barh",
    x="feature",
    y="importance",
    legend=False
)

plt.gca().invert_yaxis()
plt.title("Feature Importance - Random Forest")
plt.show()

print("\nFeature Importances:\n")
print(importance_df)

# ===============================
# 5️⃣ ROC Curve
# ===============================
RocCurveDisplay.from_predictions(y_test, y_proba)
plt.title("ROC Curve")
plt.show()

# ===============================
# 6️⃣ Precision-Recall Curve
# ===============================
PrecisionRecallDisplay.from_predictions(y_test, y_proba)
plt.title("Precision-Recall Curve")
plt.show()

# Save the model
joblib.dump(model, "header_detector.pkl")


# %%

