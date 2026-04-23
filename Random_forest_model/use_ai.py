import pandas as pd
import joblib
import openpyxl
import numpy as np


# ==============================
# UTILITAIRES FEATURES
# ==============================

def max_consecutive_empty(values):
    max_count = 0
    current = 0
    for v in values:
        if v is None:
            current += 1
            max_count = max(max_count, current)
        else:
            current = 0
    return max_count


def compute_basic_ratios(row):
    values = [cell.value for cell in row]
    non_empty = [v for v in values if v is not None]

    if not non_empty:
        return 0, 0

    str_ratio = sum(isinstance(v, str) for v in non_empty) / len(non_empty)
    num_ratio = sum(isinstance(v, (int, float)) for v in non_empty) / len(non_empty)

    return str_ratio, num_ratio


def extract_row_features_from_row(row, total_cols, ws, row_idx):

    values = [cell.value for cell in row]
    non_empty = [v for v in values if v is not None]

    if not non_empty:
        return None

    fill_ratio = len(non_empty) / total_cols
    str_ratio = sum(isinstance(v, str) for v in non_empty) / len(non_empty)
    num_ratio = sum(isinstance(v, (int, float)) for v in non_empty) / len(non_empty)

    bold_ratio = sum(
        1 for c in row if getattr(c, "font", None) and c.font.bold
    ) / total_cols

    colored_ratio = sum(
        1 for c in row
        if getattr(c, "fill", None)
        and c.fill.fgColor
        and c.fill.fgColor.rgb != "00000000"
    ) / total_cols

    str_lengths = [len(str(v)) for v in non_empty if isinstance(v, str)]
    avg_str_len = np.mean(str_lengths or [0])
    std_str_len = np.std(str_lengths or [0])

    header_keywords = [
        "date", "total", "id", "ref",
        "libellé", "libelle", "nombre", "taux",
        "montant", "appels", "décroché", "decroche",
        "période", "periode", "trimestre", "compte",
        "name", "amount",
    ]
    keyword_hits = sum(
        any(k in str(v).lower() for k in header_keywords)
        for v in non_empty if isinstance(v, str)
    )
    keyword_ratio = keyword_hits / len(non_empty)

    max_empty_streak = max_consecutive_empty(values) / total_cols
    unique_ratio = len(set(non_empty)) / len(non_empty)

    upper_ratio = sum(
        str(v).isupper() for v in non_empty if isinstance(v, str)
    ) / len(non_empty)

    special_char_ratio = sum(
        any(c in str(v) for c in ",.;:()")
        for v in non_empty if isinstance(v, str)
    ) / len(non_empty)

    num_to_str_ratio = num_ratio - str_ratio

    if row_idx < ws.max_row:
        next_row = list(ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1))[0]
        next_str_ratio, next_num_ratio = compute_basic_ratios(next_row)

        delta_str_ratio = str_ratio - next_str_ratio
        delta_num_ratio = num_ratio - next_num_ratio
    else:
        next_num_ratio = 0.0
        delta_str_ratio = 0
        delta_num_ratio = 0

    # --- contextual position features (replace row_position) ---

    prev_row_is_empty = 0.0
    if row_idx > 1:
        prev_row = list(ws.iter_rows(min_row=row_idx - 1, max_row=row_idx - 1))[0]
        prev_non_empty = [c.value for c in prev_row if c.value is not None]
        prev_row_is_empty = 1.0 if len(prev_non_empty) == 0 else 0.0

    next_row_is_numeric = 1.0 if next_num_ratio >= 0.5 else 0.0

    non_empty_rows_above = sum(
        1 for r in ws.iter_rows(min_row=1, max_row=row_idx - 1)
        if any(c.value is not None for c in r)
    )
    total_nonempty_rows = sum(
        1 for r in ws.iter_rows(min_row=1, max_row=ws.max_row)
        if any(c.value is not None for c in r)
    )
    rank_in_nonempty = (non_empty_rows_above + 1) / max(1, total_nonempty_rows)

    return [
        fill_ratio,
        str_ratio,
        num_ratio,
        bold_ratio,
        colored_ratio,
        avg_str_len,
        std_str_len,
        keyword_ratio,
        max_empty_streak,
        unique_ratio,
        upper_ratio,
        special_char_ratio,
        num_to_str_ratio,
        delta_str_ratio,
        delta_num_ratio,
        prev_row_is_empty,
        next_row_is_numeric,
        rank_in_nonempty,
    ]


# ==============================
# PREDICTION HEADER
# ==============================

def predict_header_row(filepath, model):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        total_cols = ws.max_column
        candidates = []

        for row_idx, row in enumerate(ws.iter_rows(max_row=30), start=1):
            features = extract_row_features_from_row(row, total_cols, ws, row_idx)

            if features is not None:
                proba = model.predict_proba([features])[0][1]
                candidates.append((row_idx, proba))

        if candidates:
            best_row = max(candidates, key=lambda x: x[1])
            results[sheet_name] = best_row[0]

    return results


# ==============================
# UTILISATION DU MODELE
# ==============================

model = joblib.load("header_detector.pkl")

target = target="C:/Users/docs.xlsm"

header_predictions = predict_header_row(target, model)

print("Header détectés :")
print(header_predictions)


# ==============================
# LECTURE PROPRE DU FICHIER
# ==============================

dfs = {}

for sheet_name, header_row in header_predictions.items():
    try:
        df = pd.read_excel(
            target,
            sheet_name=sheet_name,
            header=header_row - 1  # pandas commence à 0
        )
        dfs[sheet_name] = df
        print(f"✅ {sheet_name} -> header ligne {header_row}")

    except Exception as e:
        print(f"❌ Erreur lecture {sheet_name} :", e)

